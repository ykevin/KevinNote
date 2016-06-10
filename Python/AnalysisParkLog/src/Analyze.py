#!/usr/bin/python
import os
import getopt
import argparse
import sys
import json
import string

from openpyxl import Workbook
from openpyxl import load_workbook

class ParseToSaveFile(object):
    """docstring for ParseToSaveFile"""
    def __init__(self, filename, user, excel_name):
        self.filename = filename
        self.user = user
        self.excel_name = excel_name
        self.wb = Workbook(write_only = True)
        self.ws = self.wb.create_sheet()
        self.header = ['enter_license', 'leave_license', 'car_license', 'leave_time', 'actual_amount']


    def ParseLineToJson(self, line):
        sJson = line.split('INFO ')[1]
        sKey = json.loads(sJson)

        List = []
        if sKey['biz_content'].get('toll_collector_name', 0) == self.user:
            if sKey['command'] == "SET_PAYMENT_MATCH" or sKey['command'] == "SET_ABNORMAL_MATCH" or sKey['command'] == "SET_ABNORMAL_PASS_DATA":
                List.append(sKey['biz_content']['enter_car_license_number'])
                List.append(sKey['biz_content']['leave_car_license_number'])
                List.append(sKey['biz_content']['car_license_number'])
                List.append(sKey['biz_content']['leave_time'])
                List.append(sKey['biz_content']['actual_receivable'])
                return List

        return 0

    def ParseArgsToCmd(self):
        try :
            file = open(self.filename)

            while 1:
                line = file.readline();
                if not line:
                    break;

                lValue = self.ParseLineToJson(line)
                if lValue != 0 :
                    self.ws.append(lValue)
        except :
            exit

    def run(self):
        print('filename: ', self.filename,  'user: ', self.user)
        self.ws.append(self.header)
        self.ParseArgsToCmd()
        self.close()

    def close(self):
        self.wb.save(self.excel_name)

class CompareFileDiff(object):
    """docstring for CompareFileDiff"""
    def __init__(self, arg):
        self.arg = arg

    def run(self):
        print("CompareFileDiff is running")

class RepeatCharge(object):
    """docstring for RepeatCharge"""
    def __init__(self, arg):
        self.arg = arg

    def run(self):
        print("RepeatCharge is running")


def Usage():
    print("usage: --input=[value]  --output=[value]")
    print("usage: -input value -output value")
    print("type = 1 display repeat chagre analyze")
    print('type = 2 export chagre data')
    print('type = 3 compare data diff')


def main(argv):
    try:
        options, args = getopt.getopt(argv, "hc:f:t:u:o:d:s", ['help', 'compare=' 'file=', 'type=', 'user=', 'origin=', 'dst=', 'save='])
    except getopt.GetoptError:
        Usage()
        sys.exit(2)

    dArgs = {}
    for name, value in options:
        if name in ('h', '--help'):
            Usage()
        if name in ('f', '--file'):
            dArgs['file'] = value
        if name in ('t', '--type'):
            dArgs['type'] = value
        if name in ('u', '--user'):
            dArgs['user'] = value
        if name in ('c', '--compare'):
            dArgs['compare'] = value
        if name in ('o', '--orgin'):
            dArgs['origin'] = value
        if name in ('d', '--dst'):
            dArgs['dst'] = value
        if name in ('s', '--save'):
            dArgs['save'] = value

    if dArgs.get('type') == 1 :
        result = ParseToSaveFile(dArgs.get('file'), dArgs.get('user'), dArgs.get('save')+'.xlsx')
        result.run()
    elif dArgs.get('type') == 2 :
        result = CompareFileDiff()
        result.run()
    elif dArgs.get('type') == 3 :
        result = RepeatCharge()
        result.run()
    else :
        Usage()

if __name__ == "__main__":
    if len(sys.argv) > 2 :
        main(sys.argv[1:])
    else :
        Usage()
