#!/usr/bin/python
# -*- coding:utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename='009center.xlsx', use_iterators=True)
wb1 = load_workbook(filename='009.xlsx', use_iterators=True)

ws = wb.worksheets[0]
ws1 = wb1.worksheets[0]

def DiffData(license, leave_time):
    for row in ws.iter_rows():
        #print(row[1].value, row[2].value)
        if row[1].value == license and row[2].value == leave_time :
            return True
    return False

for row in ws1.iter_rows():
    #print(row[2].value, row[3].value)
    if DiffData(row[2].value, row[3].value)  == False :
        print(row[2].value, row[3].value, int(row[4].value)/100)

